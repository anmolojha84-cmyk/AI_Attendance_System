import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# ======================================
# Date
# ======================================

today = datetime.now().strftime("%d-%m-%Y")

attendance_file = os.path.join(
    "Attendance",
    f"Attendance_{today}.csv"
)

# ======================================
# Check File
# ======================================

if not os.path.exists(attendance_file):

    print("Attendance File Not Found!")
    exit()

# ======================================
# Read CSV
# ======================================

df = pd.read_csv(attendance_file)

# ======================================
# Reports Folder
# ======================================

if not os.path.exists("reports"):
    os.makedirs("reports")

excel_file = os.path.join(
    "reports",
    f"Attendance_Report_{today}.xlsx"
)

# ======================================
# Workbook
# ======================================

wb = Workbook()

ws = wb.active

ws.title = "Attendance Report"

# ======================================
# Heading
# ======================================

ws["A1"] = "AI SMART ATTENDANCE SYSTEM"
ws["A1"].font = Font(size=16, bold=True)

ws["A3"] = "Date"
ws["B3"] = today

ws["A4"] = "Total Present"
ws["B4"] = len(df)

# ======================================
# Table Header
# ======================================

ws["A6"] = "Student Name"
ws["B6"] = "Time"

ws["A6"].font = Font(bold=True)
ws["B6"].font = Font(bold=True)

# ======================================
# Data
# ======================================

row = 7

for index, data in df.iterrows():

    ws.cell(row=row, column=1).value = data["Name"]
    ws.cell(row=row, column=2).value = data["Time"]

    row += 1

# ======================================
# Save
# ======================================

wb.save(excel_file)

print("="*40)
print("Excel Report Generated Successfully")
print(excel_file)
print("="*40)